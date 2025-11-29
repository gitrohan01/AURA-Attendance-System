# attendance/utils.py
# Fixed: missing HTML import + unified export inputs + safe cleaning

from django.core.mail import EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from django.utils import timezone

from weasyprint import HTML   # <-- FIXED (this was missing)
from datetime import date, timedelta


from .models import Attendance, Session, Student, Subject
import csv, io, datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# -------------------------------------------------------
# EMAIL SENDER
# -------------------------------------------------------
def send_email_notification(subject, body, recipients):
    try:
        msg = EmailMessage(
            subject, body,
            settings.DEFAULT_FROM_EMAIL,
            recipients
        )
        msg.content_subtype = "html"
        msg.send()
    except Exception as e:
        print("Email error:", e)


# -------------------------------------------------------
# ATTENDANCE %
# -------------------------------------------------------
def attendance_percentage(student, class_group, subject, start_date, end_date):
    total = Attendance.objects.filter(
        student=student,
        session__class_group=class_group,
        session__start_time__date__range=(start_date, end_date)
    ).count()

    present = Attendance.objects.filter(
        student=student,
        session__class_group=class_group,
        present=True,
        session__start_time__date__range=(start_date, end_date)
    ).count()

    if total == 0:
        return 0

    return round((present / total) * 100, 2)


# -------------------------------------------------------
# EXPORTS — unified IDs for all
# -------------------------------------------------------

def export_class_csv(class_id):
    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["Session ID","Student ID","Student Name","Present","Verified By Face","Timestamp","Source"])

    attendances = Attendance.objects.filter(
        session__class_group_id=class_id
    ).select_related("student", "session")

    for a in attendances.order_by("timestamp"):
        w.writerow([
            a.session.session_id,
            a.student.student_id,
            f"{a.student.first_name} {a.student.last_name}",
            "Yes" if a.present else "No",
            "Yes" if a.verified_by_face else "No",
            a.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            a.source
        ])

    buf.seek(0)
    return buf



def export_session_csv(session_id):
    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["Student ID","Student Name","Present","Verified By Face","Timestamp","Source"])

    attendances = Attendance.objects.filter(
        session_id=session_id
    ).select_related("student")

    for a in attendances.order_by("timestamp"):
        w.writerow([
            a.student.student_id,
            f"{a.student.first_name} {a.student.last_name}",
            "Yes" if a.present else "No",
            "Yes" if a.verified_by_face else "No",
            a.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            a.source
        ])

    buf.seek(0)
    return buf



def export_class_xlsx(class_id):
    attendances = Attendance.objects.filter(
        session__class_group_id=class_id
    ).select_related("student", "session")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Class_{class_id}"

    headers = ["Session ID","Student ID","Student Name","Present","Verified Face","Timestamp","Source"]
    ws.append(headers)

    for a in attendances.order_by("timestamp"):
        ws.append([
            a.session.session_id,
            a.student.student_id,
            f"{a.student.first_name} {a.student.last_name}",
            "Yes" if a.present else "No",
            "Yes" if a.verified_by_face else "No",
            a.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            a.source
        ])

    # Auto column width
    for i, col in enumerate(ws.columns, 1):
        length = max(len(str(c.value)) for c in col if c.value)
        ws.column_dimensions[get_column_letter(i)].width = min(50, length + 4)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio



def export_session_xlsx(session_id):
    attendances = Attendance.objects.filter(
        session_id=session_id
    ).select_related("student")

    wb = Workbook()
    ws = wb.active
    ws.title = f"Session_{session_id}"

    headers = ["Student ID","Student Name","Present","Verified By Face","Timestamp","Source"]
    ws.append(headers)

    for a in attendances.order_by("timestamp"):
        ws.append([
            a.student.student_id,
            f"{a.student.first_name} {a.student.last_name}",
            "Yes" if a.present else "No",
            "Yes" if a.verified_by_face else "No",
            a.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            a.source
        ])

    # Auto width
    for i, col in enumerate(ws.columns, 1):
        length = max(len(str(c.value)) for c in col if c.value)
        ws.column_dimensions[get_column_letter(i)].width = min(50, length + 4)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio



def export_session_pdf(session_id):
    session = Session.objects.filter(id=session_id).first()
    attendances = Attendance.objects.filter(session=session).select_related("student")

    html = render_to_string("attendance/pdf/session_report.html", {
        "session": session,
        "attendances": attendances
    })

    pdf = HTML(string=html).write_pdf()
    return pdf



def export_subject_csv(subject_id):
    buf = io.StringIO()
    w = csv.writer(buf)

    w.writerow(["Session ID","Class","Start Time","End Time","Present","Absent"])

    sessions = Session.objects.filter(subject_id=subject_id)

    for s in sessions:
        present = Attendance.objects.filter(session=s, present=True).count()
        absent = Attendance.objects.filter(session=s, present=False).count()

        w.writerow([
            s.session_id,
            s.class_group.name,
            s.start_time,
            s.end_time or "ONGOING",
            present,
            absent
        ])

    buf.seek(0)
    return buf



def export_subject_xlsx(subject_id):
    sessions = Session.objects.filter(subject_id=subject_id)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Subject_{subject_id}"

    headers = ["Session ID","Class","Start Time","End Time","Present","Absent"]
    ws.append(headers)

    for s in sessions:
        present = Attendance.objects.filter(session=s, present=True).count()
        absent = Attendance.objects.filter(session=s, present=False).count()
        ws.append([
            s.session_id,
            s.class_group.name,
            s.start_time,
            s.end_time or "ONGOING",
            present,
            absent
        ])

    # Auto width
    for i, col in enumerate(ws.columns, 1):
        length = max(len(str(c.value)) for c in col if c.value)
        ws.column_dimensions[get_column_letter(i)].width = min(50, length + 4)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio



def export_subject_pdf(subject_input):
    # Accept either subject_id (int) or Subject object
    if isinstance(subject_input, Subject):
        subject = subject_input
    else:
        subject = Subject.objects.get(id=subject_input)

    sessions = Session.objects.filter(subject=subject)

    html = render_to_string("attendance/pdf/subject_report.html", {
        "subject": subject,
        "sessions": sessions,
    })

    bio = BytesIO()
    HTML(string=html).write_pdf(bio)
    bio.seek(0)
    return bio




# -------------------------------------------------------
# RED FLAG HELPERS
# -------------------------------------------------------
def red_flag_students_for_user(user, days=30, threshold=60):
    flagged = []

    classes = user.teacher_profile.classes.all()
    start = date.today() - timedelta(days=days)

    for c in classes:
        for s in Student.objects.filter(class_group=c):
            perc = attendance_percentage(
                s, c, None,
                start,
                date.today()
            )
            if perc < threshold:
                flagged.append({
                    "student": s,
                    "class_group": c,
                    "percentage": perc
                })

    return flagged


def render_redflag_email(student, class_group, percentage):
    subject = f"[AURA] Attendance Warning — {class_group.name}"
    body = render_to_string("attendance/email/redflag_student.html", {
        "student": student,
        "class_group": class_group,
        "percentage": percentage
    })
    return subject, body


# -------------------------------------------------------
# NEW: EMAIL EVENT BUILDERS
# -------------------------------------------------------

def build_session_start_email(session, teacher):
    html = render_to_string("attendance/email/session_start_teacher.html", {
        "teacher": teacher,
        "subject": session.subject,
        "class_group": session.class_group,
        "session_id": session.session_id,
        "start_time": session.start_time.strftime("%Y-%m-%d %H:%M"),
    })
    return "AURA — Session Started", html


def build_session_end_email(session):
    html = render_to_string("attendance/email/session_end_teacher.html", {
        "teacher": session.teacher,
        "session_id": session.session_id,
        "end_time": session.end_time.strftime("%Y-%m-%d %H:%M"),
    })
    return "AURA — Session Ended", html


def build_teacher_upload_email(teacher, class_group):
    html = render_to_string("attendance/email/class_upload_teacher.html", {
        "teacher": teacher,
        "class_group": class_group,
        "timestamp": timezone.now().strftime("%Y-%m-%d %H:%M"),
    })
    return "AURA — Attendance Uploaded", html


def build_hod_upload_email(teacher, class_group):
    html = render_to_string("attendance/email/hod_upload_notify.html", {
        "teacher": teacher,
        "class_group": class_group,
        "timestamp": timezone.now().strftime("%Y-%m-%d %H:%M"),
    })
    return "AURA — Teacher Submission", html


def build_weekly_student_report(student, class_group, present, absent, total, percentage):
    html = render_to_string("attendance/email/weekly_student_report.html", {
        "student": student,
        "class_group": class_group,
        "present": present,
        "absent": absent,
        "total": total,
        "percentage": percentage,
    })
    subject = f"AURA Weekly Attendance Report — {class_group.name}"
    return subject, html


def build_hod_weekly_digest(flagged):
    html = render_to_string("attendance/email/hod_weekly_digest.html", {
        "flagged": flagged
    })
    return "AURA — Weekly Red-Flag Digest", html

